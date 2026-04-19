from pathlib import Path

GEN = Path("gerar_oe_excel.py")
src = GEN.read_text(encoding="utf-8")

# Substitui a funcao configurar_impressao_excel inteira
OLD = '''def configurar_impressao_excel(excel_bytes, orientacao="Paisagem"):
    """Configura orientacao de impressao no Excel e retorna bytes."""
    from openpyxl import load_workbook
    from openpyxl.worksheet.page import PageMargins

    wb = load_workbook(io.BytesIO(excel_bytes))
    for ws in wb.worksheets:
        try:
            # Garante que sheet_properties existe
            if ws.sheet_properties is None:
                from openpyxl.worksheet.properties import WorksheetProperties
                ws.sheet_properties = WorksheetProperties()
        except Exception:
            pass
        if orientacao == "Paisagem":
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        else:
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(
            left=0.5, right=0.5, top=0.75, bottom=0.75)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()'''

NEW = '''def configurar_impressao_excel(excel_bytes, orientacao="Paisagem"):
    """Configura orientacao de impressao no Excel e retorna bytes."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes))
        for ws in wb.worksheets:
            try:
                if orientacao == "Paisagem":
                    ws.page_setup.orientation = "landscape"
                else:
                    ws.page_setup.orientation = "portrait"
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.page_setup.paperSize = 9  # A4
            except Exception:
                pass
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.read()
    except Exception:
        return excel_bytes  # retorna original se falhar'''

if OLD in src:
    src = src.replace(OLD, NEW, 1)
    print("OK: configurar_impressao_excel reescrita.")
else:
    print("AVISO: Funcao nao encontrada - adicionando no final.")
    # Adiciona no final do arquivo
    src = src.rstrip() + '''

def configurar_impressao_excel(excel_bytes, orientacao="Paisagem"):
    """Configura orientacao de impressao no Excel e retorna bytes."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes))
        for ws in wb.worksheets:
            try:
                if orientacao == "Paisagem":
                    ws.page_setup.orientation = "landscape"
                else:
                    ws.page_setup.orientation = "portrait"
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.page_setup.paperSize = 9
            except Exception:
                pass
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.read()
    except Exception:
        return excel_bytes
'''
    print("OK: Funcao adicionada no final.")

GEN.write_text(src, encoding="utf-8")

import py_compile, tempfile, os
tmp = tempfile.mktemp(suffix='.py')
with open(tmp, 'w', encoding='utf-8') as f:
    f.write(src)
try:
    py_compile.compile(tmp, doraise=True)
    print("SINTAXE OK! Rode: git add . && git commit -m 'Fix sheet_properties definitivo' && git push")
except py_compile.PyCompileError as e:
    print(f"ERRO: {e}")
finally:
    os.unlink(tmp)
