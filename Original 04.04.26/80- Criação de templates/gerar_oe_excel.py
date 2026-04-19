"""
gerar_oe_excel.py - Gera PDF fiel ao template Excel da OE Metalpoli
"""
from __future__ import annotations
import io
from datetime import datetime


def gerar_oe_excel(template_bytes, numero_oe, nome_cliente, itens,
                   observacoes="", config=None, logo_bytes=None):
    """Preenche template Excel com valores calculados."""
    from openpyxl import load_workbook
    if config is None:
        config = {}
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = None
    for nome_aba in wb.sheetnames:
        if 'PADR' in nome_aba.upper():
            ws = wb[nome_aba]; break
    if ws is None:
        ws = wb[wb.sheetnames[-1]]

    ws["C7"]  = config.get("nome_empresa", "Metalpoli - Fundição de Precisão")
    ws["C9"]  = config.get("endereco", "Rua Umbuzeirro Nº 74")
    ws["C11"] = config.get("bairro", "Cidade Satélite")
    ws["M11"] = config.get("cidade", "Guarulhos")
    ws["C13"] = config.get("contato", "James Machado")
    ws["M13"] = config.get("estado", "SP")
    ws["E15"] = config.get("telefone", "(11) 2954-9908")
    ws["M15"] = config.get("email", "comercial@metalpoli.com.br")
    ano = datetime.now().strftime("%y")
    ws["M5"] = f"Nº {numero_oe}"
    ws["P5"] = f"/{ano}"
    ws["C17"] = nome_cliente.upper()

    total_peso = total_qtd = total_valor = 0.0
    for i in range(14):
        ln = 21 + i
        if i < len(itens):
            it = itens[i]
            peso = float(it.get("peso_unit", 0) or 0)
            qtd  = int(it.get("qtd", 0) or 0)
            pu   = float(it.get("preco_unit", 0) or 0)
            pt   = float(it.get("preco_total", 0) or qtd * pu)
            ws[f"B{ln}"] = str(it.get("num_pedido","") or "")
            ws[f"C{ln}"] = str(it.get("num_of","") or "")
            ws[f"E{ln}"] = str(it.get("referencia","") or "")
            ws[f"F{ln}"] = str(it.get("liga","") or "")
            ws[f"G{ln}"] = str(it.get("corrida","") or "")
            ws[f"H{ln}"] = str(it.get("certificado","") or "")
            ws[f"I{ln}"] = str(it.get("cod_peca","") or "")
            ws[f"K{ln}"] = str(it.get("descricao","") or "")
            ws[f"M{ln}"] = peso
            ws[f"N{ln}"] = qtd
            ws[f"O{ln}"] = str(it.get("serie","") or "")
            ws[f"P{ln}"] = pu
            ws[f"Q{ln}"] = pt
            total_peso += peso * qtd; total_qtd += qtd; total_valor += pt
        else:
            for col in ["B","C","E","F","G","H","I","K","O"]:
                ws[f"{col}{ln}"] = None
            ws[f"M{ln}"] = None; ws[f"N{ln}"] = None
            ws[f"P{ln}"] = None; ws[f"Q{ln}"] = None

    ws["M35"] = total_peso
    ws["N35"] = int(total_qtd)
    ws["Q35"] = total_valor
    ws["B37"] = f" - {observacoes.upper()}" if observacoes else None

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out.read()


def gerar_oe_pdf(numero_oe, nome_cliente, itens, observacoes="",
                 config=None, logo_bytes=None):
    """Gera PDF paisagem fiel ao template Excel da OE."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, Image as RLImage)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    if config is None:
        config = {}

    # Paisagem A4
    pagesize = landscape(A4)
    PW, PH = pagesize
    ML = MR = 10*mm
    MT = MB = 8*mm
    W = PW - ML - MR  # ~257mm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=pagesize,
        leftMargin=ML, rightMargin=MR,
        topMargin=MT, bottomMargin=MB)

    styles = getSampleStyleSheet()
    def PS(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    BK = colors.black
    GR = colors.HexColor("#D9D9D9")

    # Dados da empresa
    nome_empresa = config.get("nome_empresa", "Metalpoli Fundição de Precisão")
    endereco = config.get("endereco", "Rua Umbuzeirro Nº 74")
    bairro   = config.get("bairro", "Cidade Satélite")
    cidade   = config.get("cidade", "Guarulhos")
    estado   = config.get("estado", "SP")
    telefone = config.get("telefone", "(11) 2954-9908")
    email    = config.get("email", "comercial@metalpoli.com.br")
    contato  = config.get("contato", "James Machado")
    ano      = datetime.now().strftime("%y")

    def fbr(v):
        return f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")
    def fpeso(v):
        return f"{v:.3f}".replace(".",",")

    story = []

    # ═══════════════════════════════════════════════════════
    # CABECALHO: Logo | Fundição de Precisão | Ordem de Entrega Nº
    # Proporcoes do Excel: B-H=logo, I-K=titulo, L-Q=OE numero
    # ═══════════════════════════════════════════════════════
    logo_w  = 55*mm   # colunas B-H
    titulo_w = 110*mm  # colunas I-K (maior)
    oe_w    = W - logo_w - titulo_w  # colunas L-Q

    if logo_bytes:
        try:
            logo_img = RLImage(io.BytesIO(logo_bytes), width=logo_w-4*mm, height=18*mm)
            logo_cell = logo_img
        except Exception:
            logo_cell = Paragraph("", PS("lc"))
    else:
        logo_cell = Paragraph("", PS("lc2"))

    cab = Table([[
        logo_cell,
        Paragraph(f"<b>Fundição de Precisão</b>",
                  PS("ft", fontSize=22, fontName="Helvetica-Bold",
                     alignment=TA_CENTER)),
        [Paragraph("<b>Ordem de Entrega</b>",
                   PS("oe1", fontSize=14, fontName="Helvetica-Bold",
                      alignment=TA_CENTER)),
         Paragraph(f"<b>Nº {numero_oe}&nbsp;&nbsp;&nbsp;&nbsp;/{ano}</b>",
                   PS("oe2", fontSize=14, fontName="Helvetica-Bold",
                      alignment=TA_CENTER))],
    ]], colWidths=[logo_w, titulo_w, oe_w])

    cab.setStyle(TableStyle([
        ("BOX",          (0,0),(-1,-1), 0.8, BK),
        ("LINEBEFORE",   (1,0),(1,0),   0.8, BK),
        ("LINEBEFORE",   (2,0),(2,0),   0.8, BK),
        ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
        ("ALIGN",        (1,0),(1,0),   "CENTER"),
        ("LEFTPADDING",  (0,0),(-1,-1), 4),
        ("RIGHTPADDING", (0,0),(-1,-1), 4),
        ("TOPPADDING",   (0,0),(-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
    ]))
    story.append(cab)

    # ═══════════════════════════════════════════════════════
    # DADOS DO FORNECEDOR - fiel ao Excel
    # Colunas: Label(B) | Valor(C-K) | Label2(L) | Valor2(M-Q)
    # ═══════════════════════════════════════════════════════
    lw  = 22*mm   # largura label esquerdo (col B)
    lw2 = 18*mm   # largura label direito (col L)
    vw2 = W - lw - lw2 - 2*mm  # valor direito (col M-Q)
    vw  = W - lw - lw2 - vw2   # valor esquerdo (col C-K)

    # Ajuste fino baseado nas proporcoes do Excel
    lw  = 22*mm
    vw  = W * 0.52
    lw2 = 16*mm
    vw2 = W - lw - vw - lw2

    def lbl(t): return Paragraph(f"<b>{t}</b>", PS("lb", fontSize=10, fontName="Helvetica-Bold"))
    def val(t): return Paragraph(str(t), PS("vl", fontSize=10, fontName="Helvetica"))
    def valb(t): return Paragraph(f"<b>{str(t)}</b>", PS("vb", fontSize=10, fontName="Helvetica-Bold"))

    dados = [
        [lbl("Fornecedor"), val(nome_empresa),        "",    ""],
        [lbl("Endereço"),   val(endereco),             "",    ""],
        [lbl("Bairro"),     val(bairro),   lbl("Cidade"), val(cidade)],
        [lbl("Contato"),    val(contato),  lbl("UF"),     val(estado)],
        [lbl("Tel."),       val(telefone), lbl("e-mail"), val(email)],
        [lbl("Cliente"),    valb(nome_cliente.upper()), "", ""],
    ]

    tbl_d = Table(dados, colWidths=[lw, vw, lw2, vw2])
    tbl_d.setStyle(TableStyle([
        ("BOX",          (0,0),(-1,-1), 0.8, BK),
        ("LINEBELOW",    (0,0),(-1,-2), 0.4, BK),
        ("LINEBEFORE",   (2,2),(2,4),   0.4, BK),
        ("SPAN",         (1,0),(3,0)),
        ("SPAN",         (1,1),(3,1)),
        ("SPAN",         (1,5),(3,5)),
        ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0),(-1,-1), 4),
        ("RIGHTPADDING", (0,0),(-1,-1), 4),
        ("TOPPADDING",   (0,0),(-1,-1), 2),
        ("BOTTOMPADDING",(0,0),(-1,-1), 2),
    ]))
    story.append(tbl_d)

    # ═══════════════════════════════════════════════════════
    # TABELA DE ITENS - proporcoes do Excel
    # B=29.1, C+D=20.3, E=29.1, F=18.1, G=15, H=21.7, I+J=49.4, K+L=41.4, M=20, N=14.4, O=22.6, P=32.2, Q=28.8
    # ═══════════════════════════════════════════════════════
    CW = [
        29*mm,   # Nº do pedido (B)
        20*mm,   # OF (C+D)
        29*mm,   # Referência (E)
        18*mm,   # Liga (F)
        15*mm,   # Corr. (G)
        22*mm,   # Certificado (H)
        49*mm,   # Codigo da Peça (I+J)
        None,    # Descrição (K+L) — calculado
        20*mm,   # Peso uni.(kg) (M)
        14*mm,   # Qtde(pçs) (N)
        23*mm,   # Série (O)
        32*mm,   # Preço Un.(R$) (P)
        None,    # Preço Total(R$) (Q) — calculado
    ]
    # Calcula colunas restantes
    fixo = sum(c for c in CW if c is not None)
    livre = W - fixo
    CW[7]  = livre * 0.55  # Descrição
    CW[12] = livre * 0.45  # Preço Total

    def ph(t): return Paragraph(t, PS("hd", fontSize=8, fontName="Helvetica-Bold",
                                       alignment=TA_CENTER, leading=9))
    def pc(t): return Paragraph(str(t), PS("cl", fontSize=8, alignment=TA_LEFT, leading=9))
    def pcc(t): return Paragraph(str(t), PS("cc", fontSize=8, alignment=TA_CENTER, leading=9))
    def pcr(t): return Paragraph(str(t), PS("cr", fontSize=8, alignment=TA_RIGHT, leading=9))

    HEADS = ["Nº do pedido","OF","Referência","Liga","Corr.","Certificado",
             "Codigo da Peça","Descrição","Peso uni.\n(kg)","Qtde\n(pçs)",
             "Série","Preço Un.\n(R$)","Preço Total\n(R$)"]

    rows = [[ph(h) for h in HEADS]]
    tp = tq = tv = 0.0

    for it in itens:
        peso = float(it.get("peso_unit",0) or 0)
        qtd  = int(it.get("qtd",0) or 0)
        pu   = float(it.get("preco_unit",0) or 0)
        pt   = float(it.get("preco_total",0) or qtd*pu)
        tp += peso*qtd; tq += qtd; tv += pt

        rows.append([
            pc(it.get("num_pedido","") or ""),
            pcc(it.get("num_of","") or ""),
            pcc(it.get("referencia","") or ""),
            pcc(it.get("liga","") or ""),
            pcc(it.get("corrida","") or ""),
            pcc(it.get("certificado","") or ""),
            pc(it.get("cod_peca","") or ""),
            pcc(it.get("descricao","") or ""),
            pcr(fpeso(peso)),
            pcc(str(qtd)),
            pcc(it.get("serie","") or ""),
            pcr(fbr(pu)),
            pcr(fbr(pt)),
        ])

    # Linhas vazias ate 14
    for _ in range(max(0, 14-len(itens))):
        rows.append([""]*13)

    # Linha total
    n = len(rows)
    rows.append([
        Paragraph("<b>Total</b>", PS("tl", fontSize=8, fontName="Helvetica-Bold")),
        "","","","","","","",
        Paragraph(f"<b>{fpeso(tp)}</b>", PS("tr", fontSize=8, fontName="Helvetica-Bold", alignment=TA_RIGHT)),
        Paragraph(f"<b>{int(tq)}</b>",   PS("tc", fontSize=8, fontName="Helvetica-Bold", alignment=TA_CENTER)),
        "",
        "",
        Paragraph(f"<b>{fbr(tv)}</b>",   PS("tv", fontSize=8, fontName="Helvetica-Bold", alignment=TA_RIGHT)),
    ])

    RH = [8*mm] + [6*mm]*(n-1) + [6*mm]
    tbl_i = Table(rows, colWidths=CW, rowHeights=RH, repeatRows=1)
    tbl_i.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),   GR),
        ("FONTNAME",      (0,0),(-1,0),   "Helvetica-Bold"),
        ("ALIGN",         (0,0),(-1,0),   "CENTER"),
        ("VALIGN",        (0,0),(-1,-1),  "MIDDLE"),
        ("GRID",          (0,0),(-1,-1),  0.5, BK),
        ("LEFTPADDING",   (0,0),(-1,-1),  2),
        ("RIGHTPADDING",  (0,0),(-1,-1),  2),
        ("TOPPADDING",    (0,0),(-1,-1),  1),
        ("BOTTOMPADDING", (0,0),(-1,-1),  1),
    ]))
    story.append(tbl_i)

    # ═══════════════════════════════════════════════════════
    # OBSERVACOES
    # ═══════════════════════════════════════════════════════
    obs_rows = [
        [Paragraph("<b>Observações</b>", PS("ot", fontSize=9, fontName="Helvetica-Bold", alignment=TA_CENTER))],
        [Paragraph(f" - {observacoes.upper()}" if observacoes else "", PS("ov", fontSize=9))],
    ]
    tbl_obs = Table(obs_rows, colWidths=[W])
    tbl_obs.setStyle(TableStyle([
        ("BOX",          (0,0),(-1,-1), 0.5, BK),
        ("LINEBELOW",    (0,0),(0,0),   0.3, BK),
        ("LEFTPADDING",  (0,0),(-1,-1), 5),
        ("TOPPADDING",   (0,0),(-1,-1), 2),
        ("BOTTOMPADDING",(0,0),(-1,-1), 2),
    ]))
    story.append(tbl_obs)

    # ═══════════════════════════════════════════════════════
    # ASSINATURAS
    # ═══════════════════════════════════════════════════════
    story.append(Spacer(1, 8*mm))
    tbl_ass = Table([[
        Paragraph("_"*45, PS("al", fontSize=9, alignment=TA_CENTER)),
        Paragraph("_"*45, PS("ar", fontSize=9, alignment=TA_CENTER)),
    ],[
        Paragraph("Carregado por:", PS("cl2", fontSize=9, alignment=TA_CENTER)),
        Paragraph("Coletado por:",  PS("cr2", fontSize=9, alignment=TA_CENTER)),
    ]], colWidths=[W/2, W/2])
    tbl_ass.setStyle(TableStyle([
        ("ALIGN",       (0,0),(-1,-1), "CENTER"),
        ("TOPPADDING",  (0,0),(-1,-1), 1),
        ("BOTTOMPADDING",(0,0),(-1,-1), 1),
    ]))
    story.append(tbl_ass)

    # ═══════════════════════════════════════════════════════
    # RODAPE - data/hora/pagina no canto inferior direito
    # ═══════════════════════════════════════════════════════
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        datetime.now().strftime("%d/%m/%Y\n%H:%M\nPg. 1"),
        PS("rd", fontSize=7, alignment=TA_RIGHT, textColor=colors.grey, leading=9)
    ))

    doc.build(story)
    buf.seek(0)
    return buf.read()


def excel_para_pdf(excel_bytes):
    import subprocess, tempfile, os
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            f.write(excel_bytes); xlsx_path = f.name
        out_dir = tempfile.mkdtemp()
        subprocess.run(['libreoffice','--headless','--convert-to','pdf',
                        '--outdir', out_dir, xlsx_path],
                       capture_output=True, timeout=30)
        pdf_path = os.path.join(out_dir,
            os.path.splitext(os.path.basename(xlsx_path))[0]+'.pdf')
        if os.path.exists(pdf_path):
            pdf = open(pdf_path,'rb').read()
            os.unlink(xlsx_path); os.unlink(pdf_path); os.rmdir(out_dir)
            return pdf
        os.unlink(xlsx_path)
        return None
    except Exception:
        return None


def configurar_impressao_excel(excel_bytes, orientacao="Paisagem"):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes))
        for ws in wb.worksheets:
            try:
                ws.page_setup.orientation = "landscape" if orientacao=="Paisagem" else "portrait"
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.page_setup.paperSize = 9
            except Exception:
                pass
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return out.read()
    except Exception:
        return excel_bytes
