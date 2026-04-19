from pathlib import Path

# Corrige o gerar_oe_excel.py - problema com nome da aba PADRAO
GEN = Path("gerar_oe_excel.py")
src = GEN.read_text(encoding="utf-8")

# Corrige a logica de selecao da aba
OLD = '''    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb["PADRAO"] if "PADRAO" in wb.sheetnames else (
         wb["PADRÃO"] if "PADRÃO" in wb.sheetnames else wb.active)'''

NEW = '''    wb = load_workbook(io.BytesIO(template_bytes))
    # Busca aba padrao por nome (tenta variantes)
    ws = None
    for nome_aba in wb.sheetnames:
        if nome_aba.upper().replace("Ã","A").replace("ã","a") in ["PADRAO","PADRÃO","PADRAO"]:
            ws = wb[nome_aba]
            break
    if ws is None:
        # Usa a ultima aba (geralmente o padrao)
        ws = wb[wb.sheetnames[-1]]'''

if OLD in src:
    src = src.replace(OLD, NEW, 1)
    print("OK: Selecao de aba corrigida.")
else:
    # Tenta versao alternativa
    OLD2 = ('    wb = load_workbook(io.BytesIO(template_bytes))\n'
            '    ws = wb["PADRAO"] if "PADRAO" in wb.sheetnames else (\n'
            '         wb["PADR\\u00c3O"] if "PADR\\u00c3O" in wb.sheetnames else wb.active)')
    if OLD2 in src:
        src = src.replace(OLD2, NEW, 1)
        print("OK: Selecao de aba corrigida (v2).")
    else:
        # Busca e substitui linha por linha
        lines = src.split('\n')
        new_lines = []
        i = 0
        fixed = False
        while i < len(lines):
            if 'load_workbook' in lines[i] and not fixed:
                new_lines.append(lines[i])
                i += 1
                # Pula linhas ate a selecao de ws
                while i < len(lines) and 'wb.active' in lines[i]:
                    i += 1
                new_lines.append('    ws = None')
                new_lines.append('    for _nm in wb.sheetnames:')
                new_lines.append('        if _nm.upper() in ["PADRAO", "PADR\\u00c3O", "PADR\\u00c1O"]:')
                new_lines.append('            ws = wb[_nm]; break')
                new_lines.append('    if ws is None: ws = wb[wb.sheetnames[-1]]')
                fixed = True
                continue
            new_lines.append(lines[i])
            i += 1
        if fixed:
            src = '\n'.join(new_lines)
            print("OK: Selecao de aba corrigida (linha a linha).")
        else:
            print("AVISO: Nao encontrado - substituindo manualmente...")
            src = src.replace(
                'wb["PADRAO"] if "PADRAO"',
                'wb[wb.sheetnames[-1]] if False'
            )

GEN.write_text(src, encoding="utf-8")

# Tambem corrige a funcao configurar_impressao_excel
OLD_CFG = '''def configurar_impressao_excel(excel_bytes, orientacao="Paisagem"):
    """Configura orientacao de impressao no Excel e retorna bytes."""
    from openpyxl import load_workbook
    from openpyxl.worksheet.page import PageMargins

    wb = load_workbook(io.BytesIO(excel_bytes))
    for ws in wb.worksheets:'''

NEW_CFG = '''def configurar_impressao_excel(excel_bytes, orientacao="Paisagem"):
    """Configura orientacao de impressao no Excel e retorna bytes."""
    from openpyxl import load_workbook
    from openpyxl.worksheet.page import PageMargins

    wb = load_workbook(io.BytesIO(excel_bytes))
    for ws in wb.worksheets:
        try:
            # Garante que sheet_properties existe
            if ws.sheet_properties is None:
                from openpyxl.worksheet.properties import WorksheetProperties
                ws.sheet_properties = WorksheetProperties()
        except Exception:
            pass'''

src = GEN.read_text(encoding="utf-8")
if OLD_CFG in src and "Garante que sheet_properties" not in src:
    src = src.replace(OLD_CFG, NEW_CFG, 1)
    GEN.write_text(src, encoding="utf-8")
    print("OK: sheet_properties corrigido.")

import py_compile, tempfile, os
tmp = tempfile.mktemp(suffix='.py')
with open(tmp, 'w', encoding='utf-8') as f:
    f.write(src)
try:
    py_compile.compile(tmp, doraise=True)
    print("SINTAXE OK! Rode: git add . && git commit -m 'Corrige sheet_properties' && git push")
except py_compile.PyCompileError as e:
    print(f"ERRO: {e}")
finally:
    os.unlink(tmp)
