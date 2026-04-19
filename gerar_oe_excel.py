"""
gerar_oe_excel.py
Gera PDF fiel ao template da OE da Metalpoli.
"""
from __future__ import annotations
import io
from datetime import datetime


def gerar_oe_excel(template_bytes, numero_oe, nome_cliente, itens,
                   observacoes="", config=None, logo_bytes=None):
    """Preenche template Excel com valores calculados."""
    from openpyxl import load_workbook
    if config is None:
        config = {}
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = None
    for nome_aba in wb.sheetnames:
        n = nome_aba.upper()
        if 'PADR' in n:
            ws = wb[nome_aba]
            break
    if ws is None:
        ws = wb[wb.sheetnames[-1]]

    ws["C7"]  = config.get("nome_empresa", "Metalpoli - Fundição de Precisão")
    ws["C9"]  = config.get("endereco", "Rua Umbuzeirro Nº 74")
    ws["C11"] = config.get("bairro", "Cidade Satélite")
    ws["M11"] = config.get("cidade", "Guarulhos")
    ws["C13"] = config.get("contato", "James Machado")
    ws["M13"] = config.get("estado", "SP")
    ws["E15"] = config.get("telefone", "(11) 2954-9908")
    ws["M15"] = config.get("email", "comercial@metalpoli.com.br")
    ano = datetime.now().strftime("%y")
    ws["M5"] = f"Nº {numero_oe}/{ano}"
    ws["P5"] = None
    ws["C17"] = nome_cliente.upper()

    total_peso = total_qtd = total_valor = 0.0
    for i in range(14):
        ln = 21 + i
        if i < len(itens):
            it = itens[i]
            peso = float(it.get("peso_unit", 0) or 0)
            qtd  = int(it.get("qtd", 0) or 0)
            pu   = float(it.get("preco_unit", 0) or 0)
            pt   = float(it.get("preco_total", 0) or qtd * pu)
            ws[f"B{ln}"] = str(it.get("num_pedido","") or "")
            ws[f"C{ln}"] = str(it.get("num_of","") or "")
            ws[f"E{ln}"] = str(it.get("referencia","") or "")
            ws[f"F{ln}"] = str(it.get("liga","") or "")
            ws[f"G{ln}"] = str(it.get("corrida","") or "")
            ws[f"H{ln}"] = str(it.get("certificado","") or "")
            ws[f"I{ln}"] = str(it.get("cod_peca","") or "")
            ws[f"K{ln}"] = str(it.get("descricao","") or "")
            ws[f"M{ln}"] = peso
            ws[f"N{ln}"] = qtd
            ws[f"O{ln}"] = str(it.get("serie","") or "")
            ws[f"P{ln}"] = pu
            ws[f"Q{ln}"] = pt
            total_peso  += peso * qtd
            total_qtd   += qtd
            total_valor += pt
        else:
            for col in ["B","C","E","F","G","H","I","K","O"]:
                ws[f"{col}{ln}"] = None
            ws[f"M{ln}"] = None; ws[f"N{ln}"] = None
            ws[f"P{ln}"] = None; ws[f"Q{ln}"] = None

    ws["M35"] = total_peso
    ws["N35"] = int(total_qtd)
    ws["Q35"] = total_valor
    if observacoes:
        ws["B37"] = f" - {observacoes.upper()}"
    else:
        ws["B37"] = None

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def gerar_oe_pdf(numero_oe, nome_cliente, itens, observacoes="",
                 config=None, logo_bytes=None):
    """Gera PDF fiel ao layout do template Excel da OE."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable, Image as RLImage)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
    from reportlab.pdfbase import pdfmetrics

    if config is None:
        config = {}

    orientacao = config.get("orientacao", "Paisagem")
    pagesize = landscape(A4) if orientacao == "Paisagem" else A4
    PW, PH = pagesize
    ML = MR = 10*mm
    MT = MB = 8*mm
    W = PW - ML - MR

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=pagesize,
        leftMargin=ML, rightMargin=MR,
        topMargin=MT, bottomMargin=MB)

    styles = getSampleStyleSheet()
    def P(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    # Cores do template
    AZUL    = colors.HexColor("#1F3864")
    BORDA   = colors.HexColor("#000000")
    CINZA_H = colors.HexColor("#D9D9D9")

    # Estilos
    s_tit_empresa = P("te", fontSize=18, fontName="Helvetica-Bold",
                       alignment=TA_CENTER)
    s_oe_titulo   = P("ot", fontSize=12, fontName="Helvetica-Bold",
                       alignment=TA_CENTER)
    s_oe_num      = P("on", fontSize=12, fontName="Helvetica-Bold",
                       alignment=TA_LEFT)
    s_label       = P("lb", fontSize=9,  fontName="Helvetica-Bold",
                       alignment=TA_LEFT)
    s_val         = P("vl", fontSize=9,  fontName="Helvetica",
                       alignment=TA_LEFT)
    s_val_b       = P("vb", fontSize=9,  fontName="Helvetica-Bold",
                       alignment=TA_LEFT)
    s_head        = P("hd", fontSize=8,  fontName="Helvetica-Bold",
                       alignment=TA_CENTER, leading=9)
    s_cell        = P("cl", fontSize=8,  fontName="Helvetica",
                       alignment=TA_LEFT, leading=9)
    s_cellc       = P("cc", fontSize=8,  fontName="Helvetica",
                       alignment=TA_CENTER, leading=9)
    s_cellr       = P("cr", fontSize=8,  fontName="Helvetica",
                       alignment=TA_RIGHT, leading=9)
    s_total_l     = P("tl", fontSize=8,  fontName="Helvetica-Bold",
                       alignment=TA_LEFT)
    s_total_r     = P("tr", fontSize=8,  fontName="Helvetica-Bold",
                       alignment=TA_RIGHT)
    s_obs_tit     = P("ob", fontSize=9,  fontName="Helvetica-Bold",
                       alignment=TA_CENTER)
    s_obs_val     = P("ov", fontSize=9,  fontName="Helvetica",
                       alignment=TA_LEFT)
    s_ass         = P("as", fontSize=9,  fontName="Helvetica",
                       alignment=TA_CENTER)
    s_footer      = P("ft", fontSize=7,  fontName="Helvetica",
                       alignment=TA_RIGHT, textColor=colors.grey)

    nome_empresa = config.get("nome_empresa", "Metalpoli Fundição de Precisão")
    endereco = config.get("endereco", "Rua Umbuzeirro Nº 74")
    bairro   = config.get("bairro", "Cidade Satélite")
    cidade   = config.get("cidade", "Guarulhos")
    estado   = config.get("estado", "SP")
    telefone = config.get("telefone", "(11) 2954-9908")
    email    = config.get("email", "comercial@metalpoli.com.br")
    contato  = config.get("contato", "James Machado")
    ano      = datetime.now().strftime("%y")

    def fmt_br(v):
        return f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")
    def fmt_peso(v):
        return f"{v:.3f}".replace(".",",")

    st = TableStyle
    TS = lambda *args: TableStyle(list(args))

    story = []

    # ═══════════════════════════════════════════════════════════════════════════
    # BLOCO 1: CABECALHO (Logo + Titulo empresa + Ordem de Entrega + Numero)
    # ═══════════════════════════════════════════════════════════════════════════
    logo_w = 45*mm
    if logo_bytes:
        try:
            logo_img = RLImage(io.BytesIO(logo_bytes), width=logo_w, height=20*mm)
            logo_cell = logo_img
        except Exception:
            logo_cell = Paragraph("", s_val)
            logo_w = 5*mm
    else:
        logo_cell = Paragraph("", s_val)
        logo_w = 5*mm

    titulo_w  = W * 0.42
    oe_w      = W - logo_w - titulo_w

    cab = Table([[
        logo_cell,
        Paragraph(nome_empresa.replace("Metalpoli ", "Metalpoli\n").replace(" Fundição", "\nFundição") if False else nome_empresa.split("Fundição de Precisão")[0] + "\n<b>Fundição de Precisão</b>" if "Fundição de Precisão" in nome_empresa else nome_empresa, 
                  P("te2", fontSize=18, fontName="Helvetica-Bold", alignment=TA_CENTER, leading=20)),
        [Paragraph("Ordem de Entrega", s_oe_titulo),
         Paragraph(f"Nº {numero_oe}&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;/{ano}", 
                   P("onum", fontSize=12, fontName="Helvetica-Bold", 
                     alignment=TA_CENTER))],
    ]], colWidths=[logo_w, titulo_w, oe_w])

    cab.setStyle(TableStyle([
        ("VALIGN",      (0,0),(-1,-1), "MIDDLE"),
        ("ALIGN",       (1,0),(1,0),   "CENTER"),
        ("LEFTPADDING", (0,0),(-1,-1), 3),
        ("RIGHTPADDING",(0,0),(-1,-1), 3),
        ("BOX",         (0,0),(-1,-1), 0.8, BORDA),
        ("LINEBEFORE",  (1,0),(1,0),   0.8, BORDA),
        ("LINEBEFORE",  (2,0),(2,0),   0.8, BORDA),
        ("TOPPADDING",  (0,0),(-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
    ]))
    story.append(cab)

    # ═══════════════════════════════════════════════════════════════════════════
    # BLOCO 2: DADOS DO FORNECEDOR
    # ═══════════════════════════════════════════════════════════════════════════
    lw = 22*mm  # largura labels
    vw = W - lw*2 - 30*mm  # largura valores principais
    cw = 15*mm  # largura cidades labels
    cv = W - lw - vw - cw  # largura valores cidade

    dados = [
        # Linha Fornecedor
        [Paragraph("<b>Fornecedor</b>", s_label),
         Paragraph(nome_empresa, s_val), "", "", "", ""],
        # Linha Endereço
        [Paragraph("<b>Endereço</b>", s_label),
         Paragraph(endereco, s_val), "", "", "", ""],
        # Linha Bairro / Cidade
        [Paragraph("<b>Bairro</b>", s_label),
         Paragraph(bairro, s_val), "",
         Paragraph("<b>Cidade</b>", s_label),
         Paragraph(cidade, s_val), ""],
        # Linha Contato / UF
        [Paragraph("<b>Contato</b>", s_label),
         Paragraph(contato, s_val), "",
         Paragraph("<b>UF</b>", s_label),
         Paragraph(estado, s_val), ""],
        # Linha Tel / email
        [Paragraph("<b>Tel.</b>", s_label),
         Paragraph(telefone, s_val), "",
         Paragraph("<b>e-mail</b>", s_label),
         Paragraph(email, s_val), ""],
        # Linha Cliente
        [Paragraph("<b>Cliente</b>", s_label),
         Paragraph(f"<b>{nome_cliente.upper()}</b>", s_val_b), "", "", "", ""],
    ]

    col_w_dados = [lw, W*0.38, W*0.06, cw, W*0.2, W*0.08]

    tbl_dados = Table(dados, colWidths=col_w_dados)
    tbl_dados.setStyle(TableStyle([
        ("BOX",          (0,0),(-1,-1), 0.8, BORDA),
        ("LINEBELOW",    (0,0),(-1,-2), 0.3, BORDA),
        ("LINEBEFORE",   (3,2),(3,4),   0.5, BORDA),
        ("SPAN",         (1,0),(5,0)),
        ("SPAN",         (1,1),(5,1)),
        ("SPAN",         (1,2),(2,2)),
        ("SPAN",         (1,3),(2,3)),
        ("SPAN",         (1,4),(2,4)),
        ("SPAN",         (4,2),(5,2)),
        ("SPAN",         (4,3),(5,3)),
        ("SPAN",         (4,4),(5,4)),
        ("SPAN",         (1,5),(5,5)),
        ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0),(-1,-1), 4),
        ("RIGHTPADDING", (0,0),(-1,-1), 4),
        ("TOPPADDING",   (0,0),(-1,-1), 2),
        ("BOTTOMPADDING",(0,0),(-1,-1), 2),
        ("BACKGROUND",   (0,0),(0,-1),  colors.white),
    ]))
    story.append(tbl_dados)

    # ═══════════════════════════════════════════════════════════════════════════
    # BLOCO 3: TABELA DE ITENS
    # ═══════════════════════════════════════════════════════════════════════════
    # Proporcoes baseadas no PDF original
    CW = [
        W*0.113,  # Nº Pedido
        W*0.065,  # OF
        W*0.085,  # Referência
        W*0.048,  # Liga
        W*0.058,  # Corr.
        W*0.073,  # Certificado
        W*0.115,  # Codigo da Peça
        W*0.085,  # Descrição
        W*0.060,  # Peso uni.(kg)
        W*0.048,  # Qtde(pçs)
        W*0.065,  # Série
        W*0.083,  # Preço Un.(R$)
        W*0.092,  # Preço Total(R$)
    ]

    HEADS = [
        "Nº do pedido", "OF", "Referência", "Liga", "Corr.",
        "Certificado", "Codigo da Peça", "Descrição",
        "Peso uni.\n(kg)", "Qtde\n(pçs)", "Série",
        "Preço Un.\n(R$)", "Preço Total\n(R$)"
    ]

    rows = [[Paragraph(h, s_head) for h in HEADS]]
    total_peso = total_qtd = total_valor = 0.0

    for it in itens:
        peso = float(it.get("peso_unit", 0) or 0)
        qtd  = int(it.get("qtd", 0) or 0)
        pu   = float(it.get("preco_unit", 0) or 0)
        pt   = float(it.get("preco_total", 0) or qtd * pu)
        total_peso  += peso * qtd
        total_qtd   += qtd
        total_valor += pt

        rows.append([
            Paragraph(str(it.get("num_pedido","") or ""), s_cell),
            Paragraph(str(it.get("num_of","") or ""), s_cellc),
            Paragraph(str(it.get("referencia","") or ""), s_cellc),
            Paragraph(str(it.get("liga","") or ""), s_cellc),
            Paragraph(str(it.get("corrida","") or ""), s_cellc),
            Paragraph(str(it.get("certificado","") or ""), s_cellc),
            Paragraph(str(it.get("cod_peca","") or ""), s_cell),
            Paragraph(str(it.get("descricao","") or ""), s_cellc),
            Paragraph(fmt_peso(peso), s_cellr),
            Paragraph(str(qtd), s_cellc),
            Paragraph(str(it.get("serie","") or ""), s_cellc),
            Paragraph(fmt_br(pu), s_cellr),
            Paragraph(fmt_br(pt), s_cellr),
        ])

    # Linhas vazias ate completar 14
    n_vazias = max(0, 14 - len(itens))
    for _ in range(n_vazias):
        rows.append([""] * 13)

    # Linha total
    n = len(rows)
    rows.append([
        Paragraph("<b>Total</b>", s_total_l),
        "","","","","","",
        "",
        Paragraph(f"<b>{fmt_peso(total_peso)}</b>", s_total_r),
        Paragraph(f"<b>{int(total_qtd)}</b>", 
                  P("trc", fontSize=8, fontName="Helvetica-Bold", alignment=TA_CENTER)),
        "",
        "",
        Paragraph(f"<b>{fmt_br(total_valor)}</b>", s_total_r),
    ])

    tbl_itens = Table(rows, colWidths=CW, repeatRows=1,
                      rowHeights=[8*mm] + [6*mm]*(len(rows)-2) + [6*mm])
    tbl_itens.setStyle(TableStyle([
        # Cabecalho
        ("BACKGROUND",    (0,0),(-1,0),    CINZA_H),
        ("FONTNAME",      (0,0),(-1,0),    "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0),    8),
        ("ALIGN",         (0,0),(-1,0),    "CENTER"),
        ("VALIGN",        (0,0),(-1,-1),   "MIDDLE"),
        # Grade
        ("GRID",          (0,0),(-1,-1),   0.5, BORDA),
        # Total
        ("BACKGROUND",    (0,n),(-1,n),    colors.white),
        ("FONTNAME",      (0,n),(-1,n),    "Helvetica-Bold"),
        # Padding
        ("LEFTPADDING",   (0,0),(-1,-1),   2),
        ("RIGHTPADDING",  (0,0),(-1,-1),   2),
        ("TOPPADDING",    (0,0),(-1,-1),   1),
        ("BOTTOMPADDING", (0,0),(-1,-1),   1),
    ]))
    story.append(tbl_itens)

    # ═══════════════════════════════════════════════════════════════════════════
    # BLOCO 4: OBSERVACOES
    # ═══════════════════════════════════════════════════════════════════════════
    obs_rows = [
        [Paragraph("<b>Observações</b>", s_obs_tit)],
        [Paragraph(f" - {observacoes.upper()}" if observacoes else "", s_obs_val)],
    ]
    tbl_obs = Table(obs_rows, colWidths=[W])
    tbl_obs.setStyle(TableStyle([
        ("BOX",          (0,0),(-1,-1), 0.5, BORDA),
        ("LINEBELOW",    (0,0),(0,0),   0.3, BORDA),
        ("LEFTPADDING",  (0,0),(-1,-1), 5),
        ("TOPPADDING",   (0,0),(-1,-1), 2),
        ("BOTTOMPADDING",(0,0),(-1,-1), 2),
        ("ALIGN",        (0,0),(0,0),   "CENTER"),
    ]))
    story.append(tbl_obs)

    # ═══════════════════════════════════════════════════════════════════════════
    # BLOCO 5: ASSINATURAS
    # ═══════════════════════════════════════════════════════════════════════════
    story.append(Spacer(1, 5*mm))
    ass_rows = [[
        Paragraph("_" * 45, s_ass),
        Paragraph("_" * 45, s_ass),
    ],[
        Paragraph("Carregado por:", s_ass),
        Paragraph("Coletado por:", s_ass),
    ]]
    tbl_ass = Table(ass_rows, colWidths=[W/2, W/2])
    tbl_ass.setStyle(TableStyle([
        ("ALIGN",        (0,0),(-1,-1), "CENTER"),
        ("TOPPADDING",   (0,0),(-1,-1), 2),
        ("BOTTOMPADDING",(0,0),(-1,-1), 2),
    ]))
    story.append(tbl_ass)

    # ═══════════════════════════════════════════════════════════════════════════
    # RODAPE com data e pagina
    # ═══════════════════════════════════════════════════════════════════════════
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        datetime.now().strftime("%d/%m/%Y\n%H:%M\nPg. 1"),
        P("rd", fontSize=7, alignment=TA_RIGHT, textColor=colors.grey, leading=9)
    ))

    doc.build(story)
    buf.seek(0)
    return buf.read()


def excel_para_pdf(excel_bytes: bytes) -> bytes | None:
    """Converte Excel para PDF usando LibreOffice."""
    import subprocess, tempfile, os
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            f.write(excel_bytes)
            xlsx_path = f.name
        out_dir = tempfile.mkdtemp()
        subprocess.run(['libreoffice','--headless','--convert-to','pdf',
                        '--outdir', out_dir, xlsx_path],
                       capture_output=True, timeout=30)
        pdf_name = os.path.splitext(os.path.basename(xlsx_path))[0] + '.pdf'
        pdf_path = os.path.join(out_dir, pdf_name)
        if os.path.exists(pdf_path):
            with open(pdf_path, 'rb') as f:
                pdf_bytes = f.read()
            os.unlink(xlsx_path); os.unlink(pdf_path); os.rmdir(out_dir)
            return pdf_bytes
        os.unlink(xlsx_path)
        return None
    except Exception:
        return None


def configurar_impressao_excel(excel_bytes, orientacao="Paisagem"):
    """Configura orientacao de impressao no Excel."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes))
        for ws in wb.worksheets:
            try:
                ws.page_setup.orientation = (
                    "landscape" if orientacao == "Paisagem" else "portrait")
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.page_setup.paperSize = 9
            except Exception:
                pass
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.read()
    except Exception:
        return excel_bytes
